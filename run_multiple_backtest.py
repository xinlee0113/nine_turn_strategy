import subprocess
import pandas as pd
import datetime
import os
import logging
import re
import sys
import argparse
import numpy as np
import matplotlib
matplotlib.use('Agg')  # 使用Agg后端，不需要GUI界面
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm
import seaborn as sns
from pathlib import Path
import platform
import time  # 添加time模块用于延迟
import traceback
import concurrent.futures  # 添加用于并行处理的模块
from matplotlib.markers import MarkerStyle  # 导入MarkerStyle

# 配置中文字体 - 在导入后立即设置
def set_chinese_font():
    # 设置显示负号
    plt.rcParams['axes.unicode_minus'] = False
    
    # 检测系统中可用的中文字体
    fonts = [f.name for f in fm.fontManager.ttflist]
    
    # 常见中文字体列表
    chinese_fonts = ['SimHei', 'Microsoft YaHei', 'Microsoft JhengHei', 'SimSun', 'NSimSun', 
                     'FangSong', 'KaiTi', 'STSong', 'STZhongsong', 'STFangsong', 'STKaiti',
                     'DengXian', 'Source Han Sans CN', 'Source Han Serif CN', 'WenQuanYi Zen Hei']
    
    # 尝试查找可用的中文字体
    font_found = False
    for font in chinese_fonts:
        if font in fonts:
            plt.rcParams['font.sans-serif'] = [font] + plt.rcParams['font.sans-serif']
            logging.info(f"找到并设置中文字体: {font}")
            font_found = True
            break
    
    # 如果没有找到中文字体，尝试从系统字体中查找包含"黑体"或"雅黑"的字体
    if not font_found:
        chinese_like_fonts = [f for f in fonts if any(keyword in f for keyword in ['黑体', '雅黑', 'Hei', 'YaHei', 'Gothic', 'Sans', '宋体', 'Song'])]
        if chinese_like_fonts:
            plt.rcParams['font.sans-serif'] = [chinese_like_fonts[0]] + plt.rcParams['font.sans-serif']
            logging.info(f"找到并设置替代中文字体: {chinese_like_fonts[0]}")
            font_found = True
    
    # 如果是Windows系统，使用更直接的字体设置方式
    if platform.system() == 'Windows' and not font_found:
        # 直接添加常见中文字体名称
        plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'SimSun'] + plt.rcParams['font.sans-serif']
        matplotlib.rc('font', family='SimHei')  # 设置默认字体
        font_found = True
        logging.info("Windows系统直接设置中文字体")
    
    if font_found:
        logging.info(f"成功设置中文字体: {plt.rcParams['font.sans-serif'][:3]}")
    else:
        logging.warning("未能找到合适的中文字体，图表中的中文可能无法正确显示")
    
    return font_found

# 立即设置中文字体
set_chinese_font()

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[logging.StreamHandler()]
)

logger = logging.getLogger(__name__)

# 定义要测试的股票符号 - 扩展为更多标的
SYMBOLS = ['QQQ', 'SPY', 'AAPL', 'TSLA', 'GOOGL', 'META', 'NVDA', 'AMZN', 'MSFT']

# 定义要测试的策略
STRATEGIES = {
    "原始策略(无做空,无费用,100万)": "--use-cache --cash 1000000",
    "原始策略(有做空,无费用,100万)": "--use-cache --enable-short --cash 1000000",
    "高级止损策略(有做空,无费用,100万)": "--use-cache --enable-short --advanced-stop-loss --cash 1000000",
    "智能止损策略(有做空,无费用,100万)": "--use-cache --enable-short --smart-stop-loss --cash 1000000",
    "原始策略(无做空,有费用,100万)": "--use-cache --real-costs --broker-type tiger --cash 1000000",
    "原始策略(有做空,有费用,100万)": "--use-cache --enable-short --real-costs --broker-type tiger --cash 1000000",
    "高级止损策略(有做空,有费用,100万)": "--use-cache --enable-short --advanced-stop-loss --real-costs --broker-type tiger --cash 1000000",
    "智能止损策略(有做空,有费用,100万)": "--use-cache --enable-short --smart-stop-loss --real-costs --broker-type tiger --cash 1000000"
}

# 设置回测天数
DAYS = 30

# 创建结果目录
OUTPUT_DIR = "out"
os.makedirs(OUTPUT_DIR, exist_ok=True)
RESULTS_DIR = os.path.join(OUTPUT_DIR, "backtest_results")
os.makedirs(RESULTS_DIR, exist_ok=True)

def parse_args():
    """解析命令行参数"""
    parser = argparse.ArgumentParser(description='运行多个回测')
    parser.add_argument('--symbols', type=str, default='QQQ', help='要测试的股票代码，用逗号分隔')
    parser.add_argument('--days', type=int, default=30, help='回测天数')
    parser.add_argument('--output-dir', type=str, default='out/backtest_results', help='输出目录')
    parser.add_argument('--batch-size', type=int, default=3, help='每批处理的股票数量，默认为3')
    parser.add_argument('--batch-wait', type=int, default=60, help='批次之间的等待时间（秒），默认为60秒')
    parser.add_argument('--combined-file', type=str, default='combined_results.xlsx', help='合并结果的文件名')
    parser.add_argument('--no-format', action='store_true', help='不格式化Excel文件')
    parser.add_argument('--no-visualize', action='store_true', help='不创建可视化图表')
    parser.add_argument('--use-cache', action='store_true', help='使用缓存数据，如果有缓存则直接使用不调用API')
    parser.add_argument('--min-wait', type=int, default=5, help='API调用之间的最小等待时间（秒）')
    parser.add_argument('--max-wait', type=int, default=15, help='API调用之间的最大等待时间（秒）')
    parser.add_argument('--max-workers', type=int, default=4, help='并行处理时的最大工作线程数')
    return parser.parse_args()

# 解析命令行参数
args = parse_args()
SYMBOLS = args.symbols
DAYS = args.days
RESULTS_DIR = args.output_dir

def check_dependencies():
    """检查所需依赖是否已安装"""
    # 检查关键依赖 - 不使用try-except，直接尝试导入
    # 如果导入失败，将直接抛出错误，程序会立即停止
    import openpyxl
    import matplotlib.pyplot
    import seaborn
    
    logger.info("所有依赖检查通过")
    return True

def get_cached_symbols(symbols, days):
    """检查哪些股票已经有缓存数据"""
    cache_dir = os.path.join("cache", "tiger")
    if not os.path.exists(cache_dir):
        return []
    
    cached_symbols = []
    for symbol in symbols:
        cache_files = [f for f in os.listdir(cache_dir) if f.startswith(f"{symbol}_") and f.endswith(".csv")]
        if cache_files:
            cached_symbols.append(symbol)
    
    return cached_symbols

def run_cache_warmup(symbols, days, min_wait=10):
    """预热缓存：为每个股票获取数据并缓存"""
    logger.info("===== 启动缓存预热阶段 =====")
    logger.info(f"预热以下股票的缓存数据: {', '.join(symbols)}")
    
    for symbol in symbols:
        if check_data_cached(symbol, days):
            logger.info(f"股票 {symbol} 已有缓存数据，跳过预热")
            continue
        
        logger.info(f"开始预热股票 {symbol} 的缓存数据...")
        cmd = f"python main.py --symbols {symbol} --days {days} --use-cache --no-plot"
        
        process = subprocess.Popen(
            cmd,
            shell=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True
        )
        
        output, error = process.communicate()
        
        if process.returncode != 0:
            logger.warning(f"预热 {symbol} 时出错，可能需要API调用，等待{min_wait}秒...")
            time.sleep(min_wait)
        else:
            logger.info(f"股票 {symbol} 缓存预热成功")
    
    logger.info("===== 缓存预热阶段完成 =====")
    
    # 返回成功缓存的股票列表
    return get_cached_symbols(symbols, days)

def check_data_cached(symbol, days):
    """检查指定股票和天数的数据是否已经缓存"""
    # 构建可能的缓存文件路径
    cache_dir = os.path.join("data", "cache")
    
    # 如果缓存目录不存在，说明没有缓存
    if not os.path.exists(cache_dir):
        logger.warning(f"缓存目录不存在: {cache_dir}")
        return False
    
    # 获取当前日期和开始日期
    end_date = datetime.datetime.now()
    start_date = end_date - datetime.timedelta(days=days)
    
    begin_str = start_date.strftime("%Y-%m-%d")
    end_str = end_date.strftime("%Y-%m-%d")
    
    # 检查该股票的缓存文件是否存在 
    # 1. 尝试精确匹配日期的缓存文件
    cache_filename = f"{cache_dir}/{symbol}_1m_{begin_str}_{end_str}.csv"
    
    if os.path.exists(cache_filename):
        file_size = os.path.getsize(cache_filename)
        logger.info(f"找到精确匹配缓存文件: {cache_filename}, 大小: {file_size} 字节")
        if file_size > 1000:  # 假设小于1KB的文件不是有效缓存
            return True
    
    # 2. 如果精确匹配不存在或太小，查找包含该股票代码的所有缓存文件
    logger.info(f"检查 {symbol} 所有可能的缓存文件...")
    all_files = [f for f in os.listdir(cache_dir) if f.startswith(f"{symbol}_1m_") and f.endswith(".csv")]
    
    if not all_files:
        logger.warning(f"未找到 {symbol} 的任何缓存文件")
        return False
    
    # 检查是否有能够覆盖所需日期范围的缓存文件
    for cache_file in all_files:
        # 尝试从文件名中提取日期范围
        try:
            # 文件命名格式：symbol_1m_YYYY-MM-DD_YYYY-MM-DD.csv
            date_parts = cache_file.replace(f"{symbol}_1m_", "").replace(".csv", "").split("_")
            if len(date_parts) == 2:
                file_begin_str, file_end_str = date_parts
                file_begin_date = datetime.datetime.strptime(file_begin_str, "%Y-%m-%d")
                file_end_date = datetime.datetime.strptime(file_end_str, "%Y-%m-%d")
                
                # 检查文件大小
                file_path = os.path.join(cache_dir, cache_file)
                file_size = os.path.getsize(file_path)
                
                # 检查日期范围是否覆盖所需日期，并且文件大小合适
                if file_begin_date <= start_date and file_end_date >= end_date and file_size > 10000:
                    logger.info(f"找到覆盖所需日期范围的缓存文件: {cache_file}, 大小: {file_size} 字节")
                    return True
                # 如果文件的结束日期足够近（如果比所需的结束日期晚或只早1-2天）
                elif file_begin_date <= start_date and (end_date - file_end_date).days <= 2 and file_size > 10000:
                    logger.info(f"找到接近所需日期范围的缓存文件: {cache_file}, 大小: {file_size} 字节")
                    return True
        except Exception as e:
            logger.warning(f"解析缓存文件名失败: {cache_file}, 错误: {e}")
            continue
    
    # 3. 检查是否有bt文件
    bt_filename = f"{cache_dir}/{symbol}_1m_bt.csv"
    if os.path.exists(bt_filename) and os.path.getsize(bt_filename) > 10000:
        logger.info(f"找到bt缓存文件: {bt_filename}, 大小: {os.path.getsize(bt_filename)} 字节")
        return True
    
    logger.warning(f"未找到合适的缓存文件: {symbol}")
    return False

def run_backtest(symbol, strategy_name, strategy_args, use_cache=False, min_wait=5, max_wait=10, skip_wait=False):
    """运行单个股票和策略的回测"""
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    log_file = f"{RESULTS_DIR}/{symbol}_{strategy_name}_{timestamp}.log"
    
    # 检查是否已缓存
    data_cached = check_data_cached(symbol, DAYS)
    
    # 构建命令参数
    cmd_args = []
    
    # 基本参数
    cmd_args.extend(["--symbols", symbol, "--days", str(DAYS), "--no-plot"])
    
    # 缓存模式处理 - 如果指定使用缓存或数据已缓存，添加--use-cache参数
    if use_cache or data_cached:
        cmd_args.append("--use-cache")
        if data_cached:
            logger.info(f"股票 {symbol} 发现缓存数据，将使用缓存")
    
    # 策略特定参数
    strategy_specific_args = strategy_args.split()
    cmd_args.extend(strategy_specific_args)
    
    # 构建最终命令
    cmd = "python main.py " + " ".join(cmd_args)
    
    # 记录命令执行信息
    if data_cached:
        logger.info(f"股票 {symbol} 的数据已缓存，使用缓存数据执行策略: {strategy_name}")
    
    logger.info(f"执行命令: {cmd}")
    
    # 添加自适应重试机制
    max_retries = 5  # 最大重试次数
    retries = 0
    consecutive_api_errors = 0  # 连续API错误计数
    success = False
    current_wait = min_wait  # 初始等待时间
    
    while retries < max_retries and not success:
        if retries > 0:
            logger.info(f"第 {retries} 次重试执行命令: {cmd}")
            
        # 执行命令
        process = subprocess.Popen(
            cmd,
            shell=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True
        )
        
        # 等待进程完成并获取输出
        output, error = process.communicate()
        
        # 处理命令返回结果
        if process.returncode != 0:
            logger.error(f"命令执行失败，错误代码: {process.returncode}")
            
            # 只显示错误的前200个字符
            error_preview = error[:200] + "..." if len(error) > 200 else error
            logger.error(f"错误信息: {error_preview}")
            
            # 检查是否是API限制错误
            if "rate limit error" in error or "Too many requests" in error:
                consecutive_api_errors += 1
                
                # 根据连续错误次数动态调整等待时间
                wait_time = min(max_wait, current_wait * (1.5 ** consecutive_api_errors))
                current_wait = wait_time
                
                logger.warning(f"检测到API调用限制错误 (连续 {consecutive_api_errors} 次)，等待{wait_time:.1f}秒后继续...")
                time.sleep(wait_time)
                retries += 1
            else:
                # 如果不是API限制错误，等待短时间后重试
                logger.warning(f"遇到非API限制错误，等待{min_wait}秒后重试...")
                time.sleep(min_wait)
                retries += 1
                # 重置连续API错误计数
                consecutive_api_errors = 0
        else:
            # 重置连续API错误计数
            consecutive_api_errors = 0
            success = True
    
    # 保存输出到日志文件
    with open(log_file, 'w', encoding='utf-8') as f:
        f.write("=== STDOUT ===\n")
        f.write(output)
        f.write("\n\n=== STDERR ===\n")
        f.write(error)
    
    logger.info(f"命令输出已保存到: {log_file}")
    
    # 分析输出日志，获取交易统计信息
    return_rate = None
    total_trades = None
    win_rate = None
    
    # 解析输出以获取关键指标
    for line in error.splitlines():
        # 检查总收益率
        if "总收益率:" in line or "总收益率：" in line:
            match = re.search(r'总收益率[：:]\s*([0-9.-]+)%', line)
            if match:
                return_rate = float(match.group(1))
                logger.info(f"找到总收益率: {return_rate}%")
                
        # 检查总交易次数
        elif "总交易次数:" in line or "总交易次数：" in line:
            match = re.search(r'总交易次数[：:]\s*([0-9]+)', line)
            if match:
                total_trades = int(match.group(1))
                logger.info(f"找到总交易次数: {total_trades}")
                
        # 检查胜率
        elif "胜率:" in line or "胜率：" in line:
            match = re.search(r'胜率[：:]\s*([0-9.-]+)%', line)
            if match:
                win_rate = float(match.group(1))
                logger.info(f"找到胜率: {win_rate}%")
    
    # 创建结果字典
    metrics = {}
    if return_rate is not None:
        metrics['收益率'] = return_rate
    if total_trades is not None:
        metrics['交易次数'] = total_trades
    if win_rate is not None:
        metrics['胜率'] = win_rate
        
    logger.info(f"提取的指标: {metrics}")
    
    # 处理等待时间逻辑
    if skip_wait:
        logger.info("跳过等待时间...")
    elif data_cached:
        # 数据已缓存，极短等待
        wait_time = min_wait / 5  # 等待时间降至最小的1/5
        logger.info(f"数据已缓存，短暂等待{wait_time:.1f}秒...")
        time.sleep(wait_time)
    else:
        # 正常等待逻辑
        wait_time = min_wait
        logger.info(f"等待{wait_time:.1f}秒，避免超过API调用限制...")
        time.sleep(wait_time)
    
    return metrics

def format_excel(excel_file, formatted_excel_file):
    """格式化Excel文件，使其更美观"""
    # 不使用try-catch，直接执行，如有错误将被抛出
    # 导入必要的模块
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.styles.differential import DifferentialStyle
    from openpyxl.formatting.rule import ColorScaleRule
    
    # 读取数据
    df = pd.read_excel(excel_file)
    
    logger.info(f"格式化Excel文件, 列名: {df.columns.tolist()}")
    
    # 检查是否是策略平均表现数据
    if "平均收益率(%)" in df.columns:
        logger.info("处理策略平均表现数据")
        # 这已经是处理过的数据，直接保存
        writer = pd.ExcelWriter(formatted_excel_file, engine='openpyxl')
        df.to_excel(writer, sheet_name='策略平均表现', index=False)
        workbook = writer.book
        ws = writer.sheets['策略平均表现']
        
        # 设置列宽
        for i, col in enumerate(df.columns):
            ws.column_dimensions[get_column_letter(i+1)].width = max(len(str(col)) * 1.5, 15)
        
        # 设置标题行格式
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # 添加条件格式 (收益率或平均收益率列)
        ws.conditional_formatting.add(
            f'B2:B{ws.max_row}',
            ColorScaleRule(start_type='min', start_color='FF9999',
                        mid_type='percentile', mid_value=50, mid_color='FFFFFF',
                        end_type='max', end_color='99FF99')
        )
        
        # 保存文件
        writer.close()
        logger.info(f"格式化Excel文件已生成: {formatted_excel_file}")
        return True
        
    # 处理详细的回测结果数据
    # 创建Excel写入器
    writer = pd.ExcelWriter(formatted_excel_file, engine='openpyxl')
    
    # 按收益率降序排列（忽略NaN值）
    df_sorted = df.sort_values(by='收益率(%)', ascending=False, na_position='last')
    
    # 准备各股票最佳策略数据
    # 过滤掉NaN值，然后对每个标的找出最高收益率的策略
    df_valid = df.dropna(subset=['收益率(%)'])
    best_indices = df_valid.groupby('股票')['收益率(%)'].idxmax()
    best_strategy_by_stock = df_valid.loc[best_indices].sort_values(by='收益率(%)', ascending=False)
    
    # 创建一个包含所有策略平均收益率的DataFrame
    strategy_avg = df_valid.groupby('策略')['收益率(%)'].mean().reset_index()
    strategy_avg = strategy_avg.sort_values(by='收益率(%)', ascending=False)
    strategy_avg.columns = ['策略', '平均收益率(%)']
    
    # 将数据写入Excel不同的sheet
    df_sorted.to_excel(writer, sheet_name='所有回测结果', index=False)
    best_strategy_by_stock.to_excel(writer, sheet_name='各股票最佳策略', index=False)
    strategy_avg.to_excel(writer, sheet_name='策略平均收益率', index=False)
    
    # 获取工作簿对象
    workbook = writer.book
    
    # 格式化所有回测结果表格
    ws = writer.sheets['所有回测结果']
    
    # 设置列宽
    for i, col in enumerate(df_sorted.columns):
        column_width = max(len(str(col)) * 1.5, df_sorted[col].astype(str).map(len).max() * 1.2)
        ws.column_dimensions[get_column_letter(i+1)].width = column_width
    
    # 设置标题行格式
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # 添加条件格式
    # 收益率列的颜色标尺
    ws.conditional_formatting.add(
        f'C2:C{len(df_sorted)+1}',
        ColorScaleRule(start_type='min', start_color='FF9999',
                    mid_type='percentile', mid_value=50, mid_color='FFFFFF',
                    end_type='max', end_color='99FF99')
    )
    
    # 胜率列的颜色标尺
    ws.conditional_formatting.add(
        f'E2:E{len(df_sorted)+1}',
        ColorScaleRule(start_type='min', start_color='FF9999',
                    mid_type='percentile', mid_value=50, mid_color='FFFFFF',
                    end_type='max', end_color='99FF99')
    )
    
    # 同样格式化其他表格
    for sheet_name in ['各股票最佳策略', '策略平均收益率']:
        ws = writer.sheets[sheet_name]
        
        # 设置列宽
        for i in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(i)].width = 15
        
        # 设置标题行格式
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # 添加条件格式 (收益率或平均收益率列)
        col_letter = 'C' if sheet_name == '各股票最佳策略' else 'B'
        ws.conditional_formatting.add(
            f'{col_letter}2:{col_letter}{ws.max_row}',
            ColorScaleRule(start_type='min', start_color='FF9999',
                        mid_type='percentile', mid_value=50, mid_color='FFFFFF',
                        end_type='max', end_color='99FF99')
        )
    
    # 保存文件
    writer.close()
    logger.info(f"格式化Excel文件已生成: {formatted_excel_file}")
    return True

def create_visualizations(excel_file, output_dir=RESULTS_DIR):
    """创建可视化图表"""
    try:
        print(f"开始创建可视化图表，Excel文件: {excel_file}, 输出目录: {output_dir}")
        logger.info(f"开始创建可视化图表，Excel文件: {excel_file}, 输出目录: {output_dir}")
        
        # 再次确保字体设置正确
        set_chinese_font()
        
        # 创建字体属性对象
        font_props = None
        
        # 设置默认字体为SimHei (黑体)
        if 'SimHei' in [f.name for f in fm.fontManager.ttflist]:
            font_props = fm.FontProperties(family='SimHei')
            logger.info("使用SimHei字体")
        elif 'Microsoft YaHei' in [f.name for f in fm.fontManager.ttflist]:
            font_props = fm.FontProperties(family='Microsoft YaHei')
            logger.info("使用Microsoft YaHei字体")
            
        # 确保文件存在
        if not os.path.exists(excel_file):
            print(f"错误: Excel文件不存在 - {excel_file}")
            logger.error(f"创建可视化图表失败: Excel文件不存在 - {excel_file}")
            return False
            
        # 确保输出目录存在
        if not os.path.exists(output_dir):
            print(f"创建输出目录: {output_dir}")
            logger.info(f"创建输出目录: {output_dir}")
            os.makedirs(output_dir)
            
        try:
            # 读取CSV文件，因为它包含更完整的数据
            csv_files = [f for f in os.listdir(output_dir) if f.startswith('backtest_results_') and f.endswith('.csv')]
            if csv_files:
                # 按修改时间排序，获取最新的CSV文件
                csv_files.sort(key=lambda x: os.path.getmtime(os.path.join(output_dir, x)), reverse=True)
                latest_csv = os.path.join(output_dir, csv_files[0])
                
                print(f"使用最新的CSV文件: {latest_csv}")
                logger.info(f"使用最新的CSV文件: {latest_csv}")
                
                # 直接读取CSV文件，不从Excel读取
                df = pd.read_csv(latest_csv)
            else:
                # 读取数据 - 检查可用的sheet名
                print(f"正在读取Excel数据，文件: {excel_file}")
                logger.info(f"正在读取Excel数据，文件: {excel_file}")
                
                # 获取可用的sheet名
                xls = pd.ExcelFile(excel_file)
                available_sheets = xls.sheet_names
                logger.info(f"Excel文件中的sheet: {available_sheets}")
                
                # 根据可用sheet选择读取方式
                if '策略平均表现' in available_sheets:
                    logger.info("检测到策略平均表现sheet，使用简化读取方式")
                    
                    # 读取策略平均表现
                    df = pd.read_excel(excel_file, sheet_name='策略平均表现')
                    
                    if '平均收益率(%)' in df.columns:
                        # 转换列名到visualization使用的列名
                        df.rename(columns={
                            '平均收益率(%)': '收益率(%)',
                            '平均交易次数': '交易次数',
                            '平均胜率(%)': '胜率(%)'
                        }, inplace=True)
                    
                    # 创建模拟的详细数据
                    df_rows = []
                    for _, row in df.iterrows():
                        df_rows.append({
                            '股票': '综合',
                            '策略': row['策略'],
                            '收益率(%)': row.get('收益率(%)', 0),
                            '交易次数': row.get('交易次数', 0),
                            '胜率(%)': row.get('胜率(%)', 0)
                        })
                    df_clean = pd.DataFrame(df_rows)
                    
                else:
                    # 使用原有的读取逻辑
                    sheet_name = '所有回测结果' if '所有回测结果' in available_sheets else available_sheets[0]
                    logger.info(f"使用sheet: {sheet_name}")
                    
                    df = pd.read_excel(excel_file, sheet_name=sheet_name)
            
            print(f"读取到的数据形状: {df.shape}")
            print(f"读取到的数据列: {df.columns.tolist()}")
            
            # 检查是否有数据可用于绘图
            print(f"检查数据是否有效: df为空: {df.empty}, 全部是NaN: {df.isna().all().all() if not df.empty else True}")
            if df.empty or df.isna().all().all():
                print("警告: 没有有效数据可用于创建可视化图表")
                logger.warning("没有有效数据可用于创建可视化图表")
                return False
            
            # 检查列名是否存在，并验证列名格式
            expected_columns = ['股票', '策略', '收益率(%)', '交易次数', '胜率(%)']
            if not all(col in df.columns for col in expected_columns[:2]):
                logger.warning(f"缺少必要的列: {[col for col in expected_columns[:2] if col not in df.columns]}")
                
                # 尝试根据列名进行推断和调整
                if '标的' in df.columns and '股票' not in df.columns:
                    df.rename(columns={'标的': '股票'}, inplace=True)
                    logger.info("将'标的'列重命名为'股票'")
            
            # 处理NaN值 - 只保留有效的行
            df_clean = df.copy()
            # 将空字符串转换为NaN
            df_clean.replace('', np.nan, inplace=True)
            
            # 只保留收益率和胜率至少有一个不为NaN的行
            df_clean = df_clean.dropna(subset=['收益率(%)'], how='all')
            
            print(f"清理NaN后的数据形状: {df_clean.shape}")
            logger.info(f"清理后的数据形状: {df_clean.shape}")
            
            if df_clean.empty:
                print("警告: 清理NaN值后没有有效数据可用于创建可视化图表")
                logger.warning("清理NaN值后没有有效数据可用于创建可视化图表")
                return False
            
            # 确保df_clean中存在所有必要的列，用默认值填充缺失的列
            for col in ['股票', '策略', '收益率(%)', '胜率(%)', '交易次数']:
                if col not in df_clean.columns:
                    df_clean[col] = np.nan if col in ['收益率(%)', '胜率(%)', '交易次数'] else '未知'
                    logger.info(f"为df_clean添加缺失的列: {col}")
            
            # 计算每个策略的平均值
            strategy_avg = df_clean.groupby('策略').agg({
                '收益率(%)': 'mean',
                '交易次数': 'mean',
                '胜率(%)': 'mean'
            }).reset_index()
            
            # 对策略平均值按收益率排序
            strategy_avg = strategy_avg.sort_values('收益率(%)', ascending=False)
            
            # 找出每个股票的最佳策略
            # 创建一个包含每个股票最佳策略的DataFrame
            best_strategies_list = []
            for stock, stock_data in df_clean.groupby('股票'):
                # 跳过收益率全部为NaN的股票
                if stock_data['收益率(%)'].isna().all():
                    continue
                
                # 找出该股票收益率最高的策略
                best_idx = stock_data['收益率(%)'].idxmax()
                best_row = stock_data.loc[best_idx].copy()
                best_strategies_list.append(best_row)
            
            # 创建DataFrame并按收益率排序
            if best_strategies_list:
                best_strategies = pd.DataFrame(best_strategies_list)
                best_strategies = best_strategies.sort_values('收益率(%)', ascending=False)
            else:
                best_strategies = pd.DataFrame()
            
            # 设置图形风格
            sns.set(style="whitegrid")
            
            try:
                # 创建一个画布，设置大小
                plt.figure(figsize=(15, 10))
                print("成功创建图形画布")
                
                # 1. 策略平均收益率对比图
                print("开始绘制策略平均收益率对比图")
                logger.info("绘制策略平均收益率对比图")
                plt.subplot(2, 2, 1)
                
                if not strategy_avg.empty and not strategy_avg['收益率(%)'].isna().all():
                    print(f"策略平均数据: {strategy_avg}")
                    # 检查是否至少有一个非NaN值
                    bars = plt.bar(strategy_avg['策略'], strategy_avg['收益率(%)'], 
                                color=sns.color_palette("viridis", len(strategy_avg)))
                    
                    # 在柱状图上显示数值
                    for bar in bars:
                        height = bar.get_height()
                        if not np.isnan(height):
                            plt.text(bar.get_x() + bar.get_width()/2., height + 0.1,
                                    f'{height:.2f}%', ha='center', va='bottom', fontsize=9,
                                    fontproperties=font_props)
                    
                    plt.title('不同策略的平均收益率对比', fontproperties=font_props)
                    plt.ylabel('平均收益率 (%)', fontproperties=font_props)
                    plt.xticks(rotation=45, ha='right', fontproperties=font_props)
                    plt.ylim(min(0, strategy_avg['收益率(%)'].min() * 1.1) if strategy_avg['收益率(%)'].min() < 0 else 0, 
                           max(0, strategy_avg['收益率(%)'].max() * 1.2))
                else:
                    print("警告: 策略平均收益率全为NaN")
                    plt.text(0.5, 0.5, '数据不足，无法创建图表', 
                           ha='center', va='center', fontsize=12, fontproperties=font_props)
                
                # 2. 各股票最佳策略收益率图
                print("开始绘制各股票最佳策略收益率图")
                logger.info("绘制各股票最佳策略收益率图")
                plt.subplot(2, 2, 2)
                
                if len(best_strategies) > 0 and not best_strategies['收益率(%)'].isna().all():
                    print(f"最佳策略数据: {best_strategies}")
                    # 检查是否至少有一个非NaN值
                    bars = plt.bar(best_strategies['股票'], best_strategies['收益率(%)'], 
                                color=sns.color_palette("muted", len(best_strategies)))
                    
                    # 在柱状图上显示数值和策略名称
                    for i, (bar, strategy) in enumerate(zip(bars, best_strategies['策略'])):
                        height = bar.get_height()
                        if not np.isnan(height):
                            plt.text(bar.get_x() + bar.get_width()/2., height + 0.5,
                                    f'{height:.2f}%', ha='center', va='bottom', fontsize=9,
                                    fontproperties=font_props)
                            plt.text(bar.get_x() + bar.get_width()/2., height/2,
                                    strategy, ha='center', va='center', fontsize=8, 
                                    color='white', rotation=90, fontproperties=font_props)
                
                    plt.title('各股票最佳策略收益率', fontproperties=font_props)
                    plt.ylabel('收益率 (%)', fontproperties=font_props)
                    plt.xticks(rotation=45, ha='right', fontproperties=font_props)
                    plt.ylim(min(0, best_strategies['收益率(%)'].min() * 1.1) if best_strategies['收益率(%)'].min() < 0 else 0, 
                           max(0, best_strategies['收益率(%)'].max() * 1.1))
                else:
                    print("警告: 无法创建各股票最佳策略收益率图表 - 数据不足")
                    plt.text(0.5, 0.5, '数据不足，无法创建图表', 
                           ha='center', va='center', fontsize=12, fontproperties=font_props)
                
                # 3. 各策略在不同股票上的收益率热力图（如果有足够的数据）
                print("开始绘制策略热力图")
                plt.subplot(2, 1, 2)
                logger.info("尝试创建策略在各股票上的收益率热力图")
                
                # 只有当有多个股票和多个策略时才创建热力图
                unique_stocks = df_clean['股票'].dropna().unique()
                unique_strategies = df_clean['策略'].dropna().unique()
                
                if len(unique_stocks) > 1 and len(unique_strategies) > 1:
                    try:
                        print(f"创建热力图数据透视表, 股票数: {len(unique_stocks)}, 策略数: {len(unique_strategies)}")
                        
                        # 创建一个临时DataFrame，确保不存在NaN值
                        df_clean_filled = df_clean.copy()
                        
                        # 创建一个透视表，行为股票，列为策略，值为收益率
                        heatmap_data = df_clean_filled.pivot_table(
                            index='股票', 
                            columns='策略', 
                            values='收益率(%)',
                            aggfunc='mean'  # 使用均值聚合
                        ).round(2)  # 四舍五入到两位小数
                        
                        print(f"热力图数据形状: {heatmap_data.shape}")
                        print(f"热力图数据: \n{heatmap_data}")
                        
                        # 绘制热力图
                        ax = sns.heatmap(heatmap_data, annot=True, cmap='RdYlGn', fmt='.2f', linewidths=0.5, 
                                cbar_kws={'label': '收益率 (%)'})
                        
                        plt.title('不同策略在各股票上的收益率表现', fontproperties=font_props)
                        # 设置坐标轴标签字体
                        plt.xlabel('策略', fontproperties=font_props)
                        plt.ylabel('股票', fontproperties=font_props)
                        
                        # 设置坐标轴刻度字体
                        plt.xticks(fontproperties=font_props, rotation=45, ha='right')
                        plt.yticks(fontproperties=font_props)
                        
                        # 使用matplotlib的colorbar添加图例
                        cbar = ax.collections[0].colorbar
                        cbar.set_label('收益率 (%)', fontproperties=font_props)
                        
                    except Exception as e:
                        print(f"创建热力图时出错: {str(e)}")
                        logger.error(f"创建热力图时出错: {e}")
                        traceback.print_exc()  # 打印完整堆栈
                        plt.text(0.5, 0.5, f'创建热力图失败: {str(e)}', 
                              ha='center', va='center', fontsize=12, fontproperties=font_props)
                else:
                    print(f"警告: 无法创建热力图 - 只有一个股票或一个策略 (股票数: {len(unique_stocks)}, 策略数: {len(unique_strategies)})")
                    plt.text(0.5, 0.5, '数据不足（需要多个股票和策略）', 
                           ha='center', va='center', fontsize=12, fontproperties=font_props)
                
                plt.tight_layout()
                print("图表布局调整完成")
                
                # 保存图像
                chart_file = f'{output_dir}/strategy_performance_visualization.png'
                print(f"正在保存图表到: {chart_file}")
                logger.info(f"保存图表到文件: {chart_file}")
                plt.savefig(chart_file, dpi=300, bbox_inches='tight')
                
                # 确认文件是否保存成功
                if os.path.exists(chart_file):
                    print(f"可视化图表已保存为: {chart_file}")
                    logger.info(f"可视化图表已保存为: {chart_file}")
                else:
                    print(f"错误: 保存图表失败，文件不存在: {chart_file}")
                    logger.error(f"保存图表失败，文件不存在: {chart_file}")
                
                # 额外创建一个胜率与收益率的散点图
                try:
                    print("开始创建胜率与收益率的散点图")
                    logger.info("创建胜率与收益率的散点图")
                    plt.figure(figsize=(10, 6))
                    
                    # 检查必要的列是否存在且有足够的非NaN值
                    required_columns = ['胜率(%)', '收益率(%)']
                    
                    # 创建只包含非NaN值的数据副本
                    scatter_df = df_clean.dropna(subset=required_columns)
                    
                    if len(scatter_df) == 0:
                        print("警告: 没有足够的非NaN数据绘制散点图")
                        plt.text(0.5, 0.5, '数据不足，无法创建散点图', 
                               ha='center', va='center', fontsize=12, fontproperties=font_props)
                    else:
                        # 确认交易次数列是否存在，否则使用默认大小
                        if '交易次数' not in scatter_df.columns or scatter_df['交易次数'].isna().all():
                            print("警告: 交易次数列不存在或全为NaN，将使用默认大小")
                            scatter_df['交易次数'] = 50  # 默认大小
                        
                        # 创建一个颜色映射，每个股票一种颜色
                        unique_stocks = scatter_df['股票'].unique()
                        colors = sns.color_palette('tab10', n_colors=len(unique_stocks))
                        color_map = {stock: color for stock, color in zip(unique_stocks, colors)}
                        
                        # 按股票和策略分组绘制
                        for i, stock in enumerate(unique_stocks):
                            stock_data = scatter_df[scatter_df['股票'] == stock]
                            
                            # 对每个策略使用不同的标记形状
                            marker_styles = ['o', 's', '^', 'D', 'v', '<', '>', 'p', '*', 'h']
                            marker_style = MarkerStyle(marker_styles[i % len(marker_styles)])  # 使用MarkerStyle创建标记
                            
                            # 绘制散点
                            plt.scatter(
                                stock_data['胜率(%)'], 
                                stock_data['收益率(%)'], 
                                s=stock_data['交易次数']/2,  # 按交易次数调整点的大小
                                c=[color_map[stock]],  # 使用预定义的颜色
                                marker=marker_style,  # 使用MarkerStyle对象
                                alpha=0.7,
                                label=f"{stock}-{stock_data['策略'].iloc[0] if not stock_data.empty else '未知'}"
                            )
                        
                        # 添加标签和图例
                        plt.title('胜率与收益率关系散点图', fontproperties=font_props)
                        plt.xlabel('胜率 (%)', fontproperties=font_props)
                        plt.ylabel('收益率 (%)', fontproperties=font_props)
                        plt.grid(True, linestyle='--', alpha=0.7)
                        
                        # 添加图例，但可能需要限制条目数量
                        if len(unique_stocks) * len(scatter_df['策略'].unique()) <= 10:
                            # 如果图例条目不多，直接显示完整图例
                            legend = plt.legend(title="股票-策略", loc="best", fontsize=8)
                        else:
                            # 如果图例条目太多，只显示股票级别的图例
                            from matplotlib.lines import Line2D
                            custom_lines = [Line2D([0], [0], color=color_map[stock], lw=4) for stock in unique_stocks]
                            legend = plt.legend(custom_lines, unique_stocks, title="股票", loc="best")
                        
                        # 设置图例标题字体
                        plt.setp(legend.get_title(), fontproperties=font_props)
                        # 设置图例文本字体
                        for text in legend.get_texts():
                            plt.setp(text, fontproperties=font_props)
                        
                        # 保存图像
                        scatter_file = f'{output_dir}/win_rate_vs_return_scatter.png'
                        print(f"正在保存散点图到: {scatter_file}")
                        logger.info(f"保存散点图到文件: {scatter_file}")
                        plt.savefig(scatter_file, dpi=300, bbox_inches='tight')
                        
                        # 确认文件是否保存成功
                        if os.path.exists(scatter_file):
                            print(f"胜率与收益率散点图已保存为: {scatter_file}")
                            logger.info(f"胜率与收益率散点图已保存为: {scatter_file}")
                        else:
                            print(f"错误: 保存散点图失败，文件不存在: {scatter_file}")
                            logger.error(f"保存散点图失败，文件不存在: {scatter_file}")
                
                except Exception as e:
                    print(f"创建散点图时出错: {str(e)}")
                    logger.error(f"创建散点图时出错: {e}")
                    traceback.print_exc()
            
            except Exception as e:
                print(f"创建图表时出错: {str(e)}")
                logger.error(f"创建图表时出错: {e}")
                traceback.print_exc()
                return False
            
            plt.close('all')  # 关闭所有图形
            print("可视化图表创建完成")
            logger.info("可视化图表创建完成")
            return True
            
        except Exception as e:
            print(f"读取Excel数据时出错: {str(e)}")
            logger.error(f"读取Excel数据或创建图表时出错: {e}")
            traceback.print_exc()
            return False
            
    except Exception as e:
        print(f"创建可视化图表的整体过程出错: {str(e)}")
        logger.error(f"创建可视化图表的整体过程出错: {e}")
        traceback.print_exc()
        return False

def save_results(df, args):
    """保存结果到CSV和Excel文件"""
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    csv_file = f"{RESULTS_DIR}/backtest_results_{timestamp}.csv"
    excel_file = f"{RESULTS_DIR}/backtest_results_{timestamp}.xlsx"
    combined_file = f"{RESULTS_DIR}/{args.combined_file}"
    
    # 保存为CSV
    df.to_csv(csv_file, index=False)
    logger.info(f"结果保存到CSV: {csv_file}")
    
    # 保存为Excel - 不使用try-except，直接执行
    df.to_excel(excel_file, index=False, sheet_name='回测结果')
    logger.info(f"结果保存到Excel: {excel_file}")
    
    # 保存合并结果
    df.to_excel(combined_file, index=False, sheet_name='回测结果')
    logger.info(f"合并结果保存到: {combined_file}")
    
    # 如果需要格式化Excel
    if not args.no_format:
        format_excel(combined_file, f"{RESULTS_DIR}/formatted_{args.combined_file}")
    
    # 如果需要生成可视化图表
    if not args.no_visualize:
        create_visualizations(f"{RESULTS_DIR}/formatted_{args.combined_file}")

def main():
    # 解析命令行参数
    args = parse_args()
    
    # 读取股票列表和策略配置
    if isinstance(SYMBOLS, str):
        symbols = SYMBOLS.split(',')
    else:
        symbols = SYMBOLS  # 已经是列表
    
    # 初始化结果表格
    results_df = pd.DataFrame()
    
    # 显示启用的参数
    if args.use_cache:
        logger.info("已启用缓存模式，将优先使用缓存数据")
    
    # 将股票列表分成批次
    symbol_batches = [symbols[i:i+args.batch_size] for i in range(0, len(symbols), args.batch_size)]
    
    logger.warning(f"注意：Tiger API有调用频率限制(每分钟最多10次)，股票将分成{len(symbol_batches)}批处理，每批{len(symbol_batches[0])}个")
    
    # 按批次处理股票
    for batch_idx, symbol_batch in enumerate(symbol_batches):
        logger.info(f"开始处理第{batch_idx+1}批股票: {', '.join(symbol_batch)}")
        
        # 批处理逻辑
        batch_success_count = 0  # 批次中成功完成的任务数
        for symbol in symbol_batch:
            logger.info(f"开始对 {symbol} 进行回测")
            for strategy_name, strategy_args in STRATEGIES.items():
                logger.info(f"运行策略: {strategy_name} 对股票: {symbol}")
                try:
                    # 执行回测
                    metrics = run_backtest(
                        symbol, 
                        strategy_name, 
                        strategy_args, 
                        use_cache=args.use_cache,
                        min_wait=args.min_wait,
                        max_wait=args.max_wait,
                        skip_wait=False  # 移除skip_wait选项，使用更安全的等待策略
                    )
                    
                    # 将结果添加到表格
                    metrics_df = pd.DataFrame({
                        '股票': [symbol],
                        '策略': [strategy_name],
                        '收益率(%)': [metrics.get('收益率', float('nan'))],
                        '交易次数': [metrics.get('交易次数', float('nan'))],
                        '胜率(%)': [metrics.get('胜率', float('nan'))]
                    })
                    results_df = pd.concat([results_df, metrics_df], ignore_index=True)
                    batch_success_count += 1
                    
                except Exception as e:
                    logger.error(f"运行回测时出错: {e}")
                    traceback.print_exc()
        
        # 批处理完成后保存阶段性结果
        if not results_df.empty and batch_idx > 0 and batch_idx % 2 == 0:
            interim_timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            interim_csv = f"{RESULTS_DIR}/interim_results_{interim_timestamp}.csv"
            logger.info(f"保存阶段性结果到: {interim_csv}")
            results_df.to_csv(interim_csv, index=False, encoding='utf-8-sig')
        
        # 批次间等待逻辑 - 简化版本
        if batch_idx < len(symbol_batches) - 1:
            # 计算待处理的股票中有多少已缓存
            next_batch = symbol_batches[batch_idx + 1]
            cached_count = sum(1 for s in next_batch if check_data_cached(s, DAYS))
            total_count = len(next_batch)
            
            # 根据缓存比例动态调整等待时间
            if total_count > 0:
                cache_ratio = cached_count / total_count
                # 缓存比例越高，等待时间越短
                adjusted_wait = args.batch_wait * (1 - cache_ratio * 0.8)
                logger.info(f"下一批中 {cached_count}/{total_count} 股票已缓存 (比例: {cache_ratio:.1%})，等待 {adjusted_wait:.1f} 秒...")
                time.sleep(adjusted_wait)
            else:
                logger.info(f"等待 {args.batch_wait} 秒后处理下一批...")
                time.sleep(args.batch_wait)
    
    # 保存结果到CSV和Excel
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    csv_file = f"{RESULTS_DIR}/backtest_results_{timestamp}.csv"
    excel_file = f"{RESULTS_DIR}/backtest_results_{timestamp}.xlsx"
    
    # 确保结果目录存在
    os.makedirs(os.path.dirname(csv_file), exist_ok=True)
    
    # 如果结果表不为空则保存
    if not results_df.empty:
        logger.info(f"保存结果到CSV: {csv_file}")
        results_df.to_csv(csv_file, index=False, encoding='utf-8-sig')
        
        logger.info(f"保存结果到Excel: {excel_file}")
        results_df.to_excel(excel_file, index=False, sheet_name='回测结果')
        
        # 保存组合结果，包括所有策略在所有股票上的平均表现
        combined_file = f"{RESULTS_DIR}/{args.combined_file}"
        logger.info(f"保存组合结果到: {combined_file}")
        
        # 按策略分组，只计算数值列的平均值
        combined_results = results_df.groupby('策略').agg({
            '收益率(%)': 'mean', 
            '交易次数': 'mean', 
            '胜率(%)': 'mean'
        }).reset_index()
        
        # 确保结果文件名统一
        combined_results.columns = ['策略', '平均收益率(%)', '平均交易次数', '平均胜率(%)']
        
        # 保存组合结果
        combined_results.to_excel(combined_file, index=False, sheet_name='平均策略表现')
        
        # 格式化Excel，使其更易读
        if not args.no_format:
            format_excel(combined_file, f"{RESULTS_DIR}/formatted_{args.combined_file}")
        
        # 创建可视化图表
        if not args.no_visualize:
            # 使用多线程处理可视化，不阻塞主流程
            import threading
            viz_thread = threading.Thread(
                target=create_visualizations,
                args=(f"{RESULTS_DIR}/formatted_{args.combined_file}", RESULTS_DIR)
            )
            viz_thread.daemon = True  # 设置为守护线程，主线程结束时自动结束
            viz_thread.start()
            logger.info("已启动后台线程创建可视化图表")
            
            # 等待可视化线程完成，但最多等待60秒
            viz_thread.join(timeout=60)
            if viz_thread.is_alive():
                logger.warning("可视化图表创建超时，但程序将继续运行")
            else:
                logger.info("可视化图表创建完成")
    else:
        logger.warning("没有结果可供保存")
        
    logger.info("回测完成")

if __name__ == "__main__":
    logger.info("脚本开始执行")
    main() # 运行所有标的股票的测试 